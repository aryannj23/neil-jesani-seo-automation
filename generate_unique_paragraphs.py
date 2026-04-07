#!/usr/bin/env python3
"""
NeilJesaniTaxResolution.com — AI Unique Paragraph Generator
============================================================
Reads NeilJesani_Programmatic_DataModel.xlsx, calls Claude API to write
unique_para_1 and unique_para_2 for every city that is still a placeholder,
then saves the results back to the workbook.

Usage:
  pip install anthropic openpyxl python-dotenv
  ANTHROPIC_API_KEY=sk-ant-... python3 generate_unique_paragraphs.py

  # Preview only (print, don't save):
  python3 generate_unique_paragraphs.py --preview

  # Process only FL + NV (Wave 1 priority):
  python3 generate_unique_paragraphs.py --wave 1

  # Resume after interruption (skips cities that already have content):
  python3 generate_unique_paragraphs.py    # re-run is safe — skips completed rows

Cost estimate: ~$0.12–0.20 total for all 98 cities at claude-haiku-4-5 pricing.
"""

import os, sys, time, argparse, logging
from pathlib import Path

try:
    import anthropic
except ImportError:
    sys.exit("ERROR: pip install anthropic")

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: pip install openpyxl")

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("generate_paragraphs.log")]
)
log = logging.getLogger(__name__)

DATA_MODEL_PATH = os.getenv("DATA_MODEL_PATH", "./NeilJesani_Programmatic_DataModel.xlsx")
API_DELAY = 0.5  # seconds between API calls

STATE_CONTEXT_NOTES = {
    "FL": "Florida has no state income tax. Key audit triggers: offshore accounts, FBAR/FATCA, Latin American holding structures, luxury real estate, and domicile change audits for recent relocatees.",
    "NV": "Nevada has no state income tax. Key audit triggers: domicile changes from CA/NY, Nevada LLC structures used by out-of-state owners, gaming/entertainment income, real estate.",
    "CA": "California has 13.3% top income tax rate. Key audit triggers: FTB concurrent audits, CA residency disputes, tech equity compensation (RSUs, ISOs, QSBS), biotech, PE/VC carried interest.",
    "NY": "New York has 10.9% state + NYC 3.876% city tax. Key audit triggers: NY residency audits, domicile disputes, hedge fund/PE income, Wall Street equity compensation, NY non-resident allocation.",
    "TX": "Texas has no state income tax. Key audit triggers: oil & gas income, S-corp structures, cryptocurrency, self-employment, and business deductions for tech sector.",
    "IL": "Illinois has 4.95% flat rate. Key audit triggers: pass-through income, real estate depreciation, business owner compensation, Chicago financial services.",
    "MA": "Massachusetts has 5% standard + 4% millionaires tax (income over $1M since 2023). Key audit triggers: biotech/pharma equity, tech RSUs, MA millionaires surtax, combined federal/state examinations.",
    "NJ": "New Jersey has 10.75% top rate. Key audit triggers: NJ domicile disputes (taxpayers claiming FL/NV residency), combined NY+NJ audit exposure, hedge fund income.",
    "CT": "Connecticut has 6.99% top rate. Key audit triggers: Fairfield County hedge fund/PE carried interest, partnership income, domicile issues, CT DRS concurrent with IRS.",
    "PA": "Pennsylvania has 3.07% flat rate + local earned income taxes. Key audit triggers: Philadelphia wage tax, PA business privilege tax, pass-through income, multi-state sourcing.",
    "CO": "Colorado has 4.40% flat rate. Key audit triggers: tech equity, cannabis industry (IRC §280E disallows deductions), real estate, Denver tech sector RSUs.",
    "WA": "Washington has 7% capital gains tax (2023+, gains over $250K). Key audit triggers: Amazon/Microsoft tech equity, WA capital gains tax compliance, business & occupation tax, crypto.",
    "NC": "North Carolina has 4.50% flat rate declining to 3.99%. Key audit triggers: Research Triangle tech/biotech equity, self-employment, multi-state sourcing for remote workers.",
    "TN": "Tennessee has no income tax (Hall Tax repealed 2021). Key audit triggers: domicile change audits from CA/NY/IL relocatees, entertainment income (Nashville), business income.",
    "MN": "Minnesota has 9.85% top rate. Key audit triggers: MN DOR aggressive audits, business income, residency changes, estate tax planning for wealthy MN families.",
    "OR": "Oregon has 9.9% + Portland Metro/Multnomah surtaxes. Key audit triggers: Portland layered income taxes, CA escapee domicile audits, OR DOR concurrent with IRS, tech equity.",
}

SYSTEM_PROMPT = """You are a senior tax resolution attorney writing content for NeilJesaniTaxResolution.com. 
Write in the voice of Neil Jesani, JD, CPA — authoritative, direct, and specific to the local market.
RULES:
- Never invent specific statistics, case outcomes, or enforcement numbers unless you flag them [VERIFY]
- Be specific to the city/market — do not write generic content that could apply to any city
- Mention specific industries, neighborhoods, or economic characteristics of the city where relevant
- Do not mention competitor firm names
- Do not use phrases like "In conclusion", "It is important to note", or other filler
- Write in active voice, no more than 3-4 sentences per paragraph within the 200-word block
- Each paragraph must be genuinely different from the other — different angle, different content
- CRITICAL CONTENT RESTRICTIONS — do NOT use any of these words or phrases:
  * "Tax Court" or "US Tax Court"
  * "litigation" or "litigate"
  * "criminal" or "Criminal Investigation"
  * "trial" or "courtroom"
  * "court-admitted" or "court proceedings"
  * "prosecute" or "prosecution"
  Instead use: "IRS dispute resolution", "IRS appeals", "tax controversy", "IRS representation", "audit defense"
- Position the firm as tax attorneys who resolve IRS problems through negotiation, appeals, and representation — NOT through court action
- Focus on: audit defense, Offers in Compromise, penalty abatement, collections defense, installment agreements, unfiled returns, IRS appeals"""

def generate_paragraphs(client: anthropic.Anthropic, city: str, state_name: str, 
                         state_abbr: str, metro: str, tax_rate: str, 
                         agency: str, nearby: str) -> tuple[str, str]:
    """Call Claude to generate both unique paragraphs for a city."""
    
    state_note = STATE_CONTEXT_NOTES.get(state_abbr, "")
    
    prompt = f"""Write TWO distinct paragraphs for the location page for tax attorney services in {city}, {state_name}.

CITY CONTEXT:
- City: {city}, {state_name} ({state_abbr})
- Metro area: {metro}
- State income tax rate: {tax_rate}
- State tax agency: {agency}
- Nearby cities: {nearby}
- State-specific context for audit risk: {state_note}

PARAGRAPH 1 — "IRS Enforcement in {city}, {state_name}: What Local Taxpayers Need to Know"
Write 200-250 words. Cover: the local tax environment, why HNWI residents in {metro} face elevated federal IRS audit risk, and what types of IRS examination are most relevant for this specific market. Be specific to {city} — mention local industries, economic characteristics, or the specific taxpayer profile in this area. Do NOT mention Tax Court, litigation, criminal investigation, or any court-related terms.

PARAGRAPH 2 — "Tax Resolution Challenges Specific to {city} Taxpayers"  
Write 200-250 words. Different angle from paragraph 1. Cover: specific tax controversy issues unique to {city} (e.g., specific industries present, income types, entity structures, or demographic characteristics that create IRS audit exposure). Mention any relevant local economic factors, recent migration patterns, or industry concentrations that affect federal tax audit risk in {city}. Do NOT mention Tax Court, litigation, criminal investigation, or any court-related terms. Focus on how a tax attorney resolves these issues through IRS negotiation, appeals, and representation.

REMINDER: Never use the words "Tax Court", "litigation", "criminal", "trial", "courtroom", "prosecute", or "court-admitted" anywhere in your response.

Format your response as:
PARA1:
[paragraph 1 text]

PARA2:
[paragraph 2 text]"""

    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt}]
    )
    
    text = response.content[0].text.strip()
    
    # Parse PARA1 / PARA2 from response
    para1, para2 = "", ""
    if "PARA1:" in text and "PARA2:" in text:
        parts = text.split("PARA2:")
        para1 = parts[0].replace("PARA1:", "").strip()
        para2 = parts[1].strip()
    else:
        # Fallback: split on double newline
        lines = [l for l in text.split("\n\n") if l.strip()]
        para1 = lines[0] if len(lines) > 0 else text
        para2 = lines[1] if len(lines) > 1 else text
    
    return para1, para2


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--preview", action="store_true", help="Print output without saving")
    parser.add_argument("--wave", default="all")
    parser.add_argument("--state", help="Process only one state (e.g. FL)")
    parser.add_argument("--city", help="Process only one city name")
    args = parser.parse_args()

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit("ERROR: ANTHROPIC_API_KEY not set. Add to .env file or environment.")
    
    client = anthropic.Anthropic(api_key=api_key)
    
    path = Path(DATA_MODEL_PATH)
    if not path.exists():
        sys.exit(f"ERROR: Data model not found at {DATA_MODEL_PATH}")
    
    wb = openpyxl.load_workbook(str(path))
    ws = wb["Location Pages — Data Model"]
    
    # Column mapping (1-indexed):
    # B=2 wave, C=3 city, D=4 state, E=5 metro, H=8 tax_rate, I=9 agency, M=13 nearby
    # N=14 unique_para_1, O=15 unique_para_2
    
    processed = 0
    skipped_complete = 0
    skipped_filter = 0
    errors = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=4), start=4):
        city_val = row[2].value  # C
        if not city_val:
            continue
        
        wave_val = str(row[1].value or "")
        state_abbr = str(row[3].value or "")
        metro = str(row[4].value or "")
        tax_rate = str(row[7].value or "")
        agency = str(row[8].value or "")
        nearby = str(row[12].value or "")
        current_para1 = row[13].value  # N
        current_para2 = row[14].value  # O

        state_names = {
            "FL":"Florida","NV":"Nevada","CA":"California","NY":"New York","TX":"Texas",
            "IL":"Illinois","MA":"Massachusetts","NJ":"New Jersey","CT":"Connecticut",
            "PA":"Pennsylvania","CO":"Colorado","WA":"Washington","NC":"North Carolina",
            "TN":"Tennessee","MN":"Minnesota","OR":"Oregon"
        }
        state_name = state_names.get(state_abbr, state_abbr)

        # Filters
        if args.city and city_val.lower() != args.city.lower():
            skipped_filter += 1
            continue
        if args.state and state_abbr.upper() != args.state.upper():
            skipped_filter += 1
            continue
        if args.wave != "all":
            if args.wave not in wave_val:
                skipped_filter += 1
                continue
        
        # Skip if already has real content (not a placeholder)
        if (current_para1 and not str(current_para1).startswith("[WRITE:") and
            current_para2 and not str(current_para2).startswith("[WRITE:")):
            log.info(f"SKIP (complete): {city_val}, {state_abbr}")
            skipped_complete += 1
            continue
        
        log.info(f"GENERATING: {city_val}, {state_abbr} ({state_name})...")
        
        try:
            para1, para2 = generate_paragraphs(
                client, city_val, state_name, state_abbr,
                metro, tax_rate, agency, nearby
            )
            
            if args.preview:
                print(f"\n{'='*60}")
                print(f"CITY: {city_val}, {state_abbr}")
                print(f"\nPARA 1:\n{para1}")
                print(f"\nPARA 2:\n{para2}")
            else:
                # Write back to Excel — row_idx is 1-based row in openpyxl
                ws.cell(row=row_idx, column=14).value = para1  # N
                ws.cell(row=row_idx, column=15).value = para2  # O
                log.info(f"  ✓ Written: {len(para1)} + {len(para2)} chars")
            
            processed += 1
            time.sleep(API_DELAY)
        
        except Exception as e:
            log.error(f"  ERROR for {city_val}: {e}")
            errors += 1
            time.sleep(2)  # Back off on error
    
    if not args.preview:
        wb.save(str(path))
        log.info(f"\nSaved: {path}")
    
    log.info(f"\n{'='*50}")
    log.info(f"COMPLETE:")
    log.info(f"  Generated: {processed}")
    log.info(f"  Skipped (already done): {skipped_complete}")
    log.info(f"  Skipped (filtered): {skipped_filter}")
    log.info(f"  Errors: {errors}")
    log.info(f"  Next step: run generate_pages.py --mode generate --wave 1 to create HTML")


if __name__ == "__main__":
    main()