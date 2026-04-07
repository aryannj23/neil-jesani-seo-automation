#!/usr/bin/env python3
"""
NeilJesaniTaxResolution.com — Programmatic Location Page Generator
================================================================
Owner: Aryan (development) | Steve (data + QA)
SOP Reference: SOP4 - Programmatic Pages

Usage:
  python generate_pages.py --mode preview --wave 1        # Preview first 5 pages (no publish)
  python generate_pages.py --mode generate --wave 1       # Generate HTML + run dedup (no publish)
  python generate_pages.py --mode publish --wave 1        # Generate + publish to WordPress
  python generate_pages.py --mode publish --wave 2        # Wave 2 publish
  python generate_pages.py --mode notices                 # Generate + publish IRS notice pages

Prerequisites:
  pip install openpyxl requests scikit-learn python-dotenv

Environment variables (create .env file — NEVER commit to git):
  WP_BASE_URL=https://neiljesanitaxresolution.com
  WP_USERNAME=your-wp-username
  WP_APP_PASSWORD=xxxx xxxx xxxx xxxx xxxx xxxx   (WordPress Application Password)
  DATA_MODEL_PATH=./NeilJesani_Programmatic_DataModel.xlsx
  TEMPLATE_PATH=./location_page_template.html
"""

import os, sys, json, time, re, argparse, logging
from pathlib import Path

# ── Optional imports with clear error messages ──────────────────────────────
try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: pip install openpyxl")

try:
    import requests
except ImportError:
    sys.exit("ERROR: pip install requests")

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    SKLEARN_AVAILABLE = True
except ImportError:
    print("WARNING: scikit-learn not installed — dedup check will be skipped.")
    print("Install: pip install scikit-learn")
    SKLEARN_AVAILABLE = False

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # .env loading is optional; env vars can be set directly

# ── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("generate_pages.log")
    ]
)
log = logging.getLogger(__name__)

# ── Config ───────────────────────────────────────────────────────────────────
WP_BASE_URL    = os.getenv("WP_BASE_URL", "https://neiljesanitaxresolution.com")
WP_USERNAME    = os.getenv("WP_USERNAME", "")
WP_APP_PASSWORD = os.getenv("WP_APP_PASSWORD", "")
DATA_MODEL_PATH = os.getenv("DATA_MODEL_PATH", "./NeilJesani_Programmatic_DataModel.xlsx")
TEMPLATE_PATH   = os.getenv("TEMPLATE_PATH", "./location_page_template.html")
DEDUP_THRESHOLD = 0.60   # Flag pages with > 60% cosine similarity
PUBLISH_DELAY   = 2      # Seconds between WP API calls (rate limiting)
CTA_PHONE       = "(800) 758-3101"

# ── State tax context sentences (injected into template) ────────────────────
STATE_TAX_CONTEXT = {
    "FL": {
        "sentence": "Florida has no state income tax, meaning federal IRS enforcement is the primary tax risk for Fort Lauderdale and Miami-area residents.",
        "detail": "Florida's lack of state income tax does not reduce federal IRS exposure. In fact, many high-net-worth individuals relocate to Florida specifically to avoid state income taxes — but this relocation can trigger IRS and home-state audits examining whether the move was bona fide. Florida also has a significant international population, creating elevated FBAR and FATCA audit risk. The Florida Department of Revenue administers sales tax and business taxes, and can conduct audits separately from the IRS."
    },
    "NV": {
        "sentence": "Nevada has no state income tax, making it a popular domicile for high-net-worth individuals — but IRS federal audit risk remains fully applicable.",
        "detail": "Nevada's absence of state income tax makes it attractive for business formation and residency changes. However, the IRS scrutinizes Nevada LLC structures used by out-of-state residents, and domicile changes to Nevada can trigger audits by the taxpayer's prior state of residence. The Nevada Department of Taxation administers commerce taxes and business license fees separately from the IRS."
    },
    "CA": {
        "sentence": "California has the highest state income tax rate in the United States at 13.3%, administered by the California Franchise Tax Board — often in parallel with IRS audits.",
        "detail": "California's 13.3% top marginal rate (plus the 1% Mental Health Services Tax) means high-income CA residents face combined federal and state rates exceeding 50% on ordinary income. The California Franchise Tax Board (FTB) conducts independent audits and frequently receives IRS audit results, which can trigger corresponding FTB examinations. CA residency and domicile issues — particularly for taxpayers who claim to have moved out of state — are an active FTB enforcement priority. Our firm handles both IRS and FTB controversy matters for California clients."
    },
    "NY": {
        "sentence": "New York imposes a top state income tax rate of 10.9%, plus New York City residents face an additional 3.876% city income tax — among the highest combined rates in the nation.",
        "detail": "New York City residents face combined federal, state, and city income tax rates that can exceed 54% on ordinary income. The New York State Department of Taxation and Finance (NYSDTF) is known for aggressive audit activity, particularly on nonresident income allocation, domicile disputes, and hedge fund/private equity carried interest. NY residency audits — examining whether taxpayers who claim to have moved to Florida or another low-tax state truly changed their domicile — are an active and expensive enforcement area."
    },
    "TX": {
        "sentence": "Texas has no state income tax, but the Texas Franchise Tax (margins tax) applies to most businesses — and IRS federal audit exposure is unchanged.",
        "detail": "Texas's lack of state income tax has made it one of the fastest-growing destinations for businesses and high-income individuals. However, the Texas Franchise Tax (margins tax) imposes an annual tax on businesses with revenues above certain thresholds, and the Texas Comptroller conducts independent audits of business tax filings. Oil and gas income, carried interest, cryptocurrency, and S-corp structures are active IRS audit targets in Texas. Federal audit exposure is the dominant concern for most Texas residents."
    },
    "IL": {
        "sentence": "Illinois imposes a flat 4.95% state income tax administered by the Illinois Department of Revenue, often in coordination with federal IRS examinations.",
        "detail": "Illinois's flat income tax rate is relatively moderate, but the state is known for aggressive business tax enforcement. The Illinois Department of Revenue conducts audits of individual and business returns, and IRS audit results are frequently shared with state authorities. Chicago-area taxpayers with pass-through business income, real estate portfolios, and equity compensation face multi-layer federal and state audit risk."
    },
    "MA": {
        "sentence": "Massachusetts imposes a 5% standard income tax rate, plus a 4% surtax on income over $1 million that took effect in 2023 — making high earners a priority for both state and federal auditors.",
        "detail": "Massachusetts's 'Millionaires Tax' (the 4% surtax on income over $1 million) took effect January 1, 2023, and is administered by the Massachusetts Department of Revenue (DOR). High-income MA residents — particularly those in technology, biotech, and financial services — now face elevated state audit risk in addition to IRS scrutiny. The MA DOR shares information with the IRS, and federal audit results can trigger corresponding state examinations."
    },
    "NJ": {
        "sentence": "New Jersey's top income tax rate of 10.75% makes it one of the highest-taxed states in the nation, with aggressive enforcement by the NJ Division of Taxation.",
        "detail": "New Jersey's combination of high income taxes, property taxes, and estate taxes makes it one of the most tax-intensive states for high-net-worth individuals. The NJ Division of Taxation actively audits residency claims by high-income individuals who claim to have moved to Florida or another low-tax state while maintaining connections to New Jersey. Combined with IRS audit risk, NJ residents face dual-layer federal and state examination exposure."
    },
    "CT": {
        "sentence": "Connecticut's top income tax rate of 6.99% applies to a high concentration of finance and hedge fund professionals in the Fairfield County corridor.",
        "detail": "Connecticut's Fairfield County — home to Greenwich, Stamford, Westport, and Darien — has one of the highest concentrations of hedge fund, private equity, and financial services professionals in the country. The Connecticut Department of Revenue Services (DRS) conducts audits that often involve carried interest, partnership income, and residency issues. CT residents are also subject to IRS scrutiny for the same income types."
    },
    "PA": {
        "sentence": "Pennsylvania's flat 3.07% state income tax is among the lower state rates, but local earned income taxes add complexity for Philadelphia and Pittsburgh area taxpayers.",
        "detail": "Pennsylvania's relatively low state income tax rate contrasts with significant local taxation: Philadelphia residents pay a 3.75% city wage tax, and most PA municipalities impose local earned income taxes through the local earned income tax (EIT) system. The Pennsylvania Department of Revenue conducts independent audits, and federal IRS audit results are shared with state authorities. Business owners face additional complexity from Pennsylvania's Business Privilege Tax and employer withholding requirements."
    },
    "CO": {
        "sentence": "Colorado's flat 4.40% state income tax is administered by the Colorado Department of Revenue, with growing audit activity targeting technology and cannabis industry income.",
        "detail": "Colorado's moderate flat income tax rate and quality of life have attracted a growing population of high-income professionals, particularly in Denver's technology sector. The Colorado Department of Revenue conducts independent audits, and IRS activity in Colorado has increased with the growth of the technology and cannabis industries. Cannabis business owners face particularly complex federal tax issues, as Section 280E of the Internal Revenue Code disallows most business deductions for cannabis businesses."
    },
    "WA": {
        "sentence": "Washington's new 7% capital gains tax on gains over $250,000 (effective 2023) creates fresh audit exposure for Amazon, Microsoft, and other tech company employees alongside ongoing federal IRS risk.",
        "detail": "Washington State's capital gains tax — imposed at 7% on capital gains over $250,000 per year — creates new state audit exposure for Washington residents with significant investment income, equity compensation, or business sale proceeds. The Washington Department of Revenue administers this tax alongside the state's business and occupation (B&O) tax. For Seattle-area tech employees with significant RSU, stock option, and investment income, both state and federal audit risk are elevated."
    },
    "NC": {
        "sentence": "North Carolina's flat income tax rate is declining under current legislation — from 4.75% in 2024 toward 3.99% by 2026 — but IRS federal audit exposure is unchanged.",
        "detail": "North Carolina's declining flat income tax rate has made it an increasingly attractive destination for businesses and individuals. The North Carolina Department of Revenue (NCDOR) conducts independent audits, and the state's Research Triangle (Raleigh-Durham) has a growing technology and life sciences sector with equity compensation complexity. IRS audit activity in NC focuses on self-employment income, business deductions, and the growing population of remote workers."
    },
    "TN": {
        "sentence": "Tennessee has no state income tax following the repeal of the Hall Tax in 2021, making it a popular relocation destination — but federal IRS audit exposure remains fully applicable.",
        "detail": "Tennessee's elimination of the Hall Tax (which previously taxed investment income) means the state now has no income tax at all. This has made Tennessee — particularly Nashville and its suburbs — one of the fastest-growing relocation destinations for high-income individuals from high-tax states. However, relocation to Tennessee does not reduce federal IRS audit risk, and the taxpayer's prior state of residence may audit whether the move was bona fide. Tennessee does impose sales tax and business taxes administered by the Tennessee Department of Revenue."
    },
    "MN": {
        "sentence": "Minnesota's top income tax rate of 9.85% is among the highest in the Midwest, and the Minnesota Department of Revenue is known for assertive audit enforcement.",
        "detail": "Minnesota's high income tax rate and assertive Department of Revenue make it a challenging tax environment for high-income individuals and business owners. The MN DOR conducts independent audits and shares information with the IRS. Minnesota also has an estate tax with a lower exemption than the federal estate tax, making estate planning and estate tax audit risk a significant consideration for high-net-worth MN residents."
    },
    "OR": {
        "sentence": "Oregon's top income tax rate of 9.9% — combined with Portland's city and metro income taxes — creates some of the highest combined income tax rates for Portland residents.",
        "detail": "Oregon residents in the Portland metro area face a layered income tax burden: Oregon state income tax (9.9% top rate), the Metro Supportive Housing Services Tax (1% on incomes over $125K), and the Multnomah County Preschool for All Tax (1.5% on incomes over $125K). Combined, these can push effective state and local rates above 12% for high earners. The Oregon Department of Revenue conducts independent audits, and federal IRS audit results are shared with state authorities. Oregon's high tax burden has increased residency change activity — often triggering both OR and IRS audits."
    }
}

# ── Schema builder ───────────────────────────────────────────────────────────
def build_location_schema(row: dict) -> str:
    """Build LocalBusiness + LegalService + FAQPage JSON-LD schema."""
    city_slug = row["city"].lower().replace(" ", "-")
    state_abbr = row["state_abbreviation"]
    
    schema = {
        "@context": "https://schema.org",
        "@graph": [
            {
                "@type": ["LegalService", "LocalBusiness"],
                "@id": f"{WP_BASE_URL}/tax-attorney-{city_slug}-{state_abbr.lower()}/#organization",
                "name": "Neil Jesani Tax Resolution",
                "description": f"IRS audit defense, tax debt resolution, and IRS dispute representation for {row['city']}, {row['state_name']} residents.",
                "url": f"{WP_BASE_URL}/tax-attorney-{city_slug}-{state_abbr.lower()}/",
                "telephone": CTA_PHONE,
                "priceRange": "Consultation: Free",
                "areaServed": {
                    "@type": "City",
                    "name": row["city"],
                    "containedInPlace": {
                        "@type": "State",
                        "name": row["state_name"]
                    }
                },
                "address": {
                    "@type": "PostalAddress",
                    "addressLocality": row["city"],
                    "addressRegion": state_abbr,
                    "addressCountry": "US"
                },
                "hasOfferCatalog": {
                    "@type": "OfferCatalog",
                    "name": "Tax Resolution Services",
                    "itemListElement": [
                        {"@type": "Offer", "itemOffered": {"@type": "Service", "name": "IRS Audit Defense"}},
                        {"@type": "Offer", "itemOffered": {"@type": "Service", "name": "IRS Appeals Representation"}},
                        {"@type": "Offer", "itemOffered": {"@type": "Service", "name": "IRS Collections Defense"}},
                        {"@type": "Offer", "itemOffered": {"@type": "Service", "name": "Offer in Compromise"}},
                        {"@type": "Offer", "itemOffered": {"@type": "Service", "name": "IRS Penalty Abatement"}},
                        {"@type": "Offer", "itemOffered": {"@type": "Service", "name": "Unfiled Tax Returns"}},
                    ]
                },
                "sameAs": [
                    "https://neiljesani.com",
                    "https://www.avvo.com/tax-attorney/",
                ],
                "founder": {
                    "@type": "Person",
                    "name": "Neil Jesani",
                    "jobTitle": "Tax Attorney, CPA",
                    "knowsAbout": ["IRS Audit Defense", "Tax Controversy", "IRS Appeals", "Tax Debt Resolution"]
                }
            },
            {
                "@type": "FAQPage",
                "mainEntity": [
                    {
                        "@type": "Question",
                        "name": f"Do I need to visit your office in Miami or Las Vegas to get help with my IRS issue in {row['city']}?",
                        "acceptedAnswer": {
                            "@type": "Answer",
                            "text": f"No. Our firm represents clients remotely throughout {row['state_name']} and all 50 states. IRS matters are handled by phone, video conference, and secure document portal. We file Form 2848 so the IRS communicates with us directly. Call {CTA_PHONE} to schedule a confidential consultation from {row['city']}."
                        }
                    },
                    {
                        "@type": "Question",
                        "name": f"What IRS field office handles tax matters for {row['city']}, {row['state_name']} taxpayers?",
                        "acceptedAnswer": {
                            "@type": "Answer",
                            "text": f"The nearest IRS field office serving {row['city']} is located at {row['local_irs_office_address']}, phone {row['local_irs_phone']}. However, do not contact the IRS directly — retain representation first. Once we file Form 2848, all IRS communications come to us."
                        }
                    },
                    {
                        "@type": "Question",
                        "name": f"Can I challenge an IRS decision from {row['city']}, {row['state_name']}?",
                        "acceptedAnswer": {
                            "@type": "Answer",
                            "text": f"Yes. If you disagree with an IRS determination, you have the right to appeal through the IRS Office of Appeals. Our tax attorneys represent {row['city']}-area clients through the full appeals process. The deadline to respond to a Notice of Deficiency is 90 days — acting quickly is critical."
                        }
                    }
                ]
            },
            {
                "@type": "BreadcrumbList",
                "itemListElement": [
                    {"@type": "ListItem", "position": 1, "name": "Home", "item": f"{WP_BASE_URL}/"},
                    {"@type": "ListItem", "position": 2, "name": "Locations", "item": f"{WP_BASE_URL}/locations/"},
                    {"@type": "ListItem", "position": 3, "name": f"{row['city']}, {state_abbr}", "item": f"{WP_BASE_URL}/tax-attorney-{city_slug}-{state_abbr.lower()}/"}
                ]
            }
        ]
    }
    return json.dumps(schema, indent=2)


def build_notice_schema(notice: dict) -> str:
    """Build FAQPage + HowTo schema for IRS notice pages."""
    schema = {
        "@context": "https://schema.org",
        "@graph": [
            {
                "@type": "FAQPage",
                "mainEntity": [
                    {
                        "@type": "Question",
                        "name": f"What is IRS Notice {notice['code']}?",
                        "acceptedAnswer": {
                            "@type": "Answer",
                            "text": f"IRS Notice {notice['code']} is the {notice['name']}. {notice['meaning']} You have {notice['deadline']} to respond."
                        }
                    },
                    {
                        "@type": "Question",
                        "name": f"What should I do when I receive IRS {notice['code']}?",
                        "acceptedAnswer": {
                            "@type": "Answer",
                            "text": f"{notice['action']}. Do not ignore this notice — {notice['penalty']}. Call {CTA_PHONE} for a confidential consultation."
                        }
                    }
                ]
            },
            {
                "@type": "LegalService",
                "name": "Neil Jesani Tax Resolution",
                "description": f"IRS {notice['code']} {notice['name']} response and defense services.",
                "telephone": CTA_PHONE,
                "url": f"{WP_BASE_URL}/irs-notice/{notice['code'].lower()}/"
            }
        ]
    }
    return json.dumps(schema, indent=2)


# ── Data loader ──────────────────────────────────────────────────────────────
def load_location_data(path: str, wave_filter: str = None) -> list[dict]:
    """Load location page rows from Excel data model."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Location Pages — Data Model"]
    
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[2]:  # No city = skip
            continue
        
        num, wave, city, state_abbr, metro, irs_addr, irs_phone, tax_rate, \
        agency, agency_url, pop, income, nearby, unique1, unique2, court, \
        primary_kw, vol, priority, gen_status, pub_status = row[:21]
        
        # Skip if unique paragraphs are still placeholders
        if unique1 and unique1.startswith("[WRITE:"):
            if wave_filter:  # In non-preview mode, warn but continue
                log.warning(f"INCOMPLETE: {city}, {state_abbr} — unique_para_1 not written. Page will be SKIPPED.")
            continue
        
        wave_str = str(wave or "")
        if wave_filter and wave_filter not in wave_str:
            continue
        
        state_ctx = STATE_TAX_CONTEXT.get(state_abbr, {})
        
        rows.append({
            "city": str(city),
            "state_abbreviation": str(state_abbr),
            "state_name": get_state_name(state_abbr),
            "metro_area": str(metro or ""),
            "local_irs_office_address": str(irs_addr or "[VERIFY: IRS.gov]"),
            "local_irs_phone": str(irs_phone or "[VERIFY: IRS.gov]"),
            "state_income_tax_rate": str(tax_rate or ""),
            "state_tax_agency_name": str(agency or ""),
            "state_tax_agency_url": str(agency_url or ""),
            "population_est": str(pop or ""),
            "median_hhi_est": str(income or ""),
            "nearby_cities": str(nearby or ""),
            "unique_local_paragraph_1": str(unique1 or ""),
            "unique_local_paragraph_2": str(unique2 or ""),
            "local_court_info": str(court or ""),
            "state_tax_context_sentence": state_ctx.get("sentence", ""),
            "state_tax_context_detail": state_ctx.get("detail", ""),
            "cta_phone": CTA_PHONE,
            "wave": wave_str,
            "priority": str(priority or ""),
        })
    
    log.info(f"Loaded {len(rows)} location rows (wave filter: {wave_filter or 'all'})")
    return rows


def load_notice_data(path: str) -> list[dict]:
    """Load IRS notice page rows from Excel data model."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["IRS Notice Pages — Data Model"]
    
    notices = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        num, code, name, meaning, deadline, action, reference, penalty, hnwi, priority, gen_st, pub_st = row[:12]
        notices.append({
            "code": str(code),
            "name": str(name),
            "meaning": str(meaning),
            "deadline": str(deadline),
            "action": str(action),
            "reference": str(reference),
            "penalty": str(penalty),
            "hnwi": str(hnwi),
            "priority": str(priority),
        })
    
    log.info(f"Loaded {len(notices)} notice rows")
    return notices


STATE_NAMES = {
    "FL":"Florida","NV":"Nevada","CA":"California","NY":"New York","TX":"Texas",
    "IL":"Illinois","MA":"Massachusetts","NJ":"New Jersey","CT":"Connecticut",
    "PA":"Pennsylvania","CO":"Colorado","WA":"Washington","NC":"North Carolina",
    "TN":"Tennessee","MN":"Minnesota","OR":"Oregon"
}

def get_state_name(abbr: str) -> str:
    return STATE_NAMES.get(abbr, abbr)


# ── Template injector ────────────────────────────────────────────────────────
def render_location_page(template: str, row: dict) -> str:
    """Inject all variables into the location page template."""
    # Build nearby cities as HTML list items with internal links
    nearby_links = ""
    for city in [c.strip() for c in row["nearby_cities"].split(",") if c.strip()]:
        city_slug = city.lower().replace(" ", "-").replace(".", "")
        state_abbr = row["state_abbreviation"].lower()
        nearby_links += f'    <li><a href="/tax-attorney-{city_slug}-{state_abbr}/">{city}, {row["state_abbreviation"]} Tax Attorney</a></li>\n'
    
    schema_json = build_location_schema(row)
    
    content = template
    content = content.replace("{{schema_json}}", schema_json)
    content = content.replace("{{nearby_cities_links}}", nearby_links)
    
    for key, val in row.items():
        content = content.replace(f"{{{{{key}}}}}", str(val))
    
    return content


def render_notice_page(notice: dict) -> str:
    """Generate IRS notice page HTML."""
    schema_json = build_notice_schema(notice)
    
    return f"""<script type="application/ld+json">
{schema_json}
</script>

<nav class="njtr-breadcrumb">
  <ol>
    <li><a href="/">Home</a></li>
    <li><a href="/irs-notices/">IRS Notices</a></li>
    <li>IRS {notice['code']}</li>
  </ol>
</nav>

<section class="njtr-hero">
  <h1>IRS Notice {notice['code']}: {notice['name']} — What It Means and How to Respond</h1>
  
  <div class="njtr-answer-lead">
    <p>IRS Notice {notice['code']} — the {notice['name']} — means: {notice['meaning']} You have <strong>{notice['deadline']}</strong> to respond. The recommended action: {notice['action']}. If you ignore this notice, the consequences are: {notice['penalty']}. This notice is issued under {notice['reference']}. Contact our team at <strong>{CTA_PHONE}</strong> for immediate confidential assistance.</p>
  </div>

  <div class="njtr-tldr-box">
    <strong>TL;DR — IRS {notice['code']}</strong>
    <ul>
      <li><strong>What it is:</strong> {notice['name']}</li>
      <li><strong>What it means:</strong> {notice['meaning']}</li>
      <li><strong>Your deadline:</strong> {notice['deadline']}</li>
      <li><strong>What to do:</strong> {notice['action']}</li>
      <li><strong>Consequence of ignoring:</strong> {notice['penalty']}</li>
    </ul>
  </div>

  <div class="njtr-cta-block">
    <strong>Received IRS {notice['code']}? We Can Help.</strong>
    <p>Neil Jesani's team of tax attorneys, CPAs, and Enrolled Agents responds to IRS {notice['code']} notices for high-net-worth individuals and business owners nationwide.</p>
    <a href="/contact/" class="njtr-cta-btn">Schedule a Free Consultation</a>
    <span class="njtr-cta-phone">or Call {CTA_PHONE} — Confidential, No Obligation</span>
  </div>
</section>

<section class="njtr-section">
  <h2>What Is IRS Notice {notice['code']}?</h2>
  <p>IRS Notice {notice['code']} is formally called the <strong>{notice['name']}</strong>. {notice['meaning']}</p>
  <p>This notice is issued under <strong>{notice['reference']}</strong>.</p>
  <table>
    <thead><tr><th>Item</th><th>Detail</th></tr></thead>
    <tbody>
      <tr><td>Notice Name</td><td>{notice['name']}</td></tr>
      <tr><td>Response Deadline</td><td>{notice['deadline']}</td></tr>
      <tr><td>IRS Authority</td><td>{notice['reference']}</td></tr>
      <tr><td>Consequence of Non-Response</td><td>{notice['penalty']}</td></tr>
      <tr><td>HNWI Relevance</td><td>{notice['hnwi']}</td></tr>
    </tbody>
  </table>
</section>

<section class="njtr-section">
  <h2>How to Respond to IRS {notice['code']}: Step-by-Step</h2>
  <ol>
    <li><strong>Do not ignore the notice.</strong> {notice['penalty']}</li>
    <li><strong>Retain professional representation immediately.</strong> Do not contact the IRS directly without counsel. File Form 2848 to redirect all IRS communications to your representative.</li>
    <li><strong>Review the notice carefully.</strong> Identify what the IRS is claiming, the amount at issue, and the exact response deadline.</li>
    <li><strong>Gather your documentation.</strong> Collect records that address the specific items the IRS is questioning.</li>
    <li><strong>Respond within {notice['deadline']}.</strong> {notice['action']}</li>
    <li><strong>Follow up.</strong> Confirm receipt and track the status of your response.</li>
  </ol>
</section>

<div class="njtr-cta-block">
  <strong>IRS {notice['code']} — Act Within Your Deadline.</strong>
  <p>Our tax attorneys and CPAs respond to IRS {notice['code']} notices every day. Confidential consultation — no obligation.</p>
  <a href="/contact/" class="njtr-cta-btn">Schedule a Free Consultation</a>
  <span class="njtr-cta-phone">{CTA_PHONE}</span>
</div>

<section class="njtr-faq" itemscope itemtype="https://schema.org/FAQPage">
  <h2>Frequently Asked Questions — IRS {notice['code']}</h2>

  <div itemscope itemprop="mainEntity" itemtype="https://schema.org/Question">
    <h3 itemprop="name">What is IRS Notice {notice['code']}?</h3>
    <div itemscope itemprop="acceptedAnswer" itemtype="https://schema.org/Answer">
      <p itemprop="text">IRS Notice {notice['code']} — the {notice['name']} — means {notice['meaning']} It is issued under {notice['reference']}.</p>
    </div>
  </div>

  <div itemscope itemprop="mainEntity" itemtype="https://schema.org/Question">
    <h3 itemprop="name">How long do I have to respond to IRS {notice['code']}?</h3>
    <div itemscope itemprop="acceptedAnswer" itemtype="https://schema.org/Answer">
      <p itemprop="text">Your response deadline is {notice['deadline']}. Missing this deadline has serious consequences: {notice['penalty']}. If you need an extension, contact your representative immediately — do not ignore the deadline or wait.</p>
    </div>
  </div>

  <div itemscope itemprop="mainEntity" itemtype="https://schema.org/Question">
    <h3 itemprop="name">What happens if I ignore IRS {notice['code']}?</h3>
    <div itemscope itemprop="acceptedAnswer" itemtype="https://schema.org/Answer">
      <p itemprop="text">If you ignore IRS {notice['code']}: {notice['penalty']}. Ignoring IRS notices does not make them go away — it almost always results in significantly worse outcomes than responding professionally and on time. Call {CTA_PHONE} immediately if you have received this notice and are unsure how to respond.</p>
    </div>
  </div>

  <div itemscope itemprop="mainEntity" itemtype="https://schema.org/Question">
    <h3 itemprop="name">Do I need a tax attorney to respond to IRS {notice['code']}?</h3>
    <div itemscope itemprop="acceptedAnswer" itemtype="https://schema.org/Answer">
      <p itemprop="text">For complex returns, high-income taxpayers, business owners, or any situation where the amount at issue is significant, professional representation is strongly recommended. A tax attorney or CPA can file Form 2848 to redirect all IRS communications to their office, develop the most effective response strategy, and prevent the most common and costly mistakes that self-represented taxpayers make when responding to IRS notices. Call {CTA_PHONE} for a free confidential assessment.</p>
    </div>
  </div>

</section>

<div class="njtr-cta-block njtr-cta-final">
  <strong>Ready to Resolve Your IRS {notice['code']} Notice?</strong>
  <p>Neil Jesani's team handles IRS {notice['code']} responses for clients nationwide. Free confidential consultation.</p>
  <a href="/contact/" class="njtr-cta-btn">Schedule Your Free Consultation</a>
  <span class="njtr-cta-phone">{CTA_PHONE}</span>
</div>
"""


# ── Deduplication ─────────────────────────────────────────────────────────────
def run_dedup_check(pages: list[dict], threshold: float = DEDUP_THRESHOLD) -> list[tuple]:
    """
    Run TF-IDF cosine similarity dedup check.
    Returns list of (slug_a, slug_b, similarity_score) for flagged pairs.
    """
    if not SKLEARN_AVAILABLE:
        log.warning("Dedup check skipped — scikit-learn not installed.")
        return []
    
    texts = [p["content"] for p in pages]
    slugs = [p["slug"] for p in pages]
    
    log.info(f"Running dedup check on {len(texts)} pages...")
    vectorizer = TfidfVectorizer(
        max_features=5000,
        stop_words="english",
        ngram_range=(1, 2)
    )
    tfidf_matrix = vectorizer.fit_transform(texts)
    sim_matrix = cosine_similarity(tfidf_matrix)
    
    flagged = []
    for i in range(len(slugs)):
        for j in range(i + 1, len(slugs)):
            score = sim_matrix[i][j]
            if score > threshold:
                flagged.append((slugs[i], slugs[j], round(score, 3)))
    
    if flagged:
        log.warning(f"DEDUP: {len(flagged)} page pairs exceed {threshold:.0%} similarity threshold:")
        for a, b, score in sorted(flagged, key=lambda x: x[2], reverse=True):
            log.warning(f"  {score:.1%} — {a} vs {b}")
    else:
        log.info(f"DEDUP: All pages within {threshold:.0%} threshold ✓")
    
    return flagged


# ── WordPress publisher ──────────────────────────────────────────────────────
class WordPressPublisher:
    def __init__(self):
        self.base_url = WP_BASE_URL.rstrip("/")
        self.session = requests.Session()
        if not WP_USERNAME or not WP_APP_PASSWORD:
            log.warning("WP credentials not set — publish mode will fail. Set WP_USERNAME and WP_APP_PASSWORD in .env")
        self.session.auth = (WP_USERNAME, WP_APP_PASSWORD)
        self.session.headers.update({"Content-Type": "application/json"})
    
    def _post_exists(self, slug: str) -> bool:
        """Check if a post with this slug already exists."""
        resp = self.session.get(
            f"{self.base_url}/wp-json/wp/v2/pages",
            params={"slug": slug, "per_page": 1}
        )
        if resp.status_code == 200:
            return len(resp.json()) > 0
        return False
    
    def publish_page(self, slug: str, title: str, content: str, 
                     meta_description: str, categories: list = None) -> dict:
        """Publish a single page via WordPress REST API."""
        if self._post_exists(slug):
            log.info(f"SKIP (exists): /{slug}/")
            return {"skipped": True, "slug": slug}
        
        payload = {
            "title": title,
            "content": content,
            "slug": slug,
            "status": "publish",
            "type": "page",
            "meta": {
                # Rank Math SEO fields
                "rank_math_description": meta_description,
                "rank_math_focus_keyword": slug.replace("-", " ").replace("tax attorney ", ""),
                "rank_math_robots": ["index", "follow"],
            }
        }
        
        resp = self.session.post(
            f"{self.base_url}/wp-json/wp/v2/pages",
            data=json.dumps(payload)
        )
        
        if resp.status_code in (200, 201):
            page_id = resp.json().get("id")
            log.info(f"PUBLISHED: /{slug}/ (ID: {page_id})")
            return {"success": True, "slug": slug, "id": page_id}
        else:
            log.error(f"FAILED: /{slug}/ — {resp.status_code}: {resp.text[:200]}")
            return {"success": False, "slug": slug, "error": resp.text[:200]}
    
    def submit_to_gsc(self, urls: list) -> None:
        """Note: Google Search Console indexing API requires OAuth — log URLs for manual submission."""
        log.info(f"GSC: Submit these {len(urls)} URLs to Google Search Console for indexing:")
        for url in urls:
            log.info(f"  {url}")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="NeilJesaniTaxResolution.com — Programmatic Page Generator")
    parser.add_argument("--mode", choices=["preview","generate","publish","notices"],
                        default="preview", help="Execution mode")
    parser.add_argument("--wave", default="1",
                        help="Location page wave to process")
    parser.add_argument("--output-dir", default="./generated_pages",
                        help="Directory for generated HTML files")
    args = parser.parse_args()

    # Load template
    template_path = Path(TEMPLATE_PATH)
    if not template_path.exists():
        sys.exit(f"ERROR: Template not found at {TEMPLATE_PATH}")
    template = template_path.read_text(encoding="utf-8")
    
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── NOTICES MODE ─────────────────────────────────────────────────────────
    if args.mode == "notices":
        notices = load_notice_data(DATA_MODEL_PATH)
        publisher = WordPressPublisher() if args.mode == "publish" else None
        
        generated = []
        for notice in notices:
            slug = f"irs-notice/{notice['code'].lower()}"
            title = f"IRS Notice {notice['code']}: {notice['name']} — What It Means & How to Respond"
            content = render_notice_page(notice)
            
            out_path = output_dir / f"notice_{notice['code']}.html"
            out_path.write_text(content, encoding="utf-8")
            log.info(f"GENERATED: {out_path.name}")
            generated.append({"slug": slug, "content": content, "title": title})
        
        log.info(f"\n✓ Generated {len(generated)} IRS notice pages in {output_dir}/")
        return

    # ── LOCATION MODES ───────────────────────────────────────────────────────
    wave_filter = None if args.wave == "all" else args.wave
    rows = load_location_data(DATA_MODEL_PATH, wave_filter)
    
    if not rows:
        log.warning("No rows with completed unique paragraphs found. Add content to unique_para_1 and unique_para_2 columns first.")
        sys.exit(0)
    
    if args.mode == "preview":
        rows = rows[:5]
        log.info(f"PREVIEW MODE: processing first {len(rows)} rows only.")
    
    # Generate pages
    generated = []
    for row in rows:
        city_slug = row["city"].lower().replace(" ", "-").replace(".", "")
        state_slug = row["state_abbreviation"].lower()
        slug = f"tax-attorney-{city_slug}-{state_slug}"
        
        title = f"Tax Attorney in {row['city']}, {row['state_name']}: IRS Audit & Tax Resolution Help"
        meta_desc = (f"Facing the IRS in {row['city']}? Neil Jesani's team of tax attorneys, CPAs, "
                     f"and Enrolled Agents helps {row['city']} residents resolve IRS audits, tax debt, "
                     f"and collections. Free consultation: {CTA_PHONE}.")
        
        content = render_location_page(template, row)
        
        out_path = output_dir / f"{slug}.html"
        out_path.write_text(content, encoding="utf-8")
        
        generated.append({
            "slug": slug,
            "title": title,
            "content": content,
            "meta_description": meta_desc,
            "url": f"{WP_BASE_URL}/{slug}/"
        })
    
    log.info(f"Generated {len(generated)} HTML files to {output_dir}/")
    
    # Dedup check
    flagged = run_dedup_check(generated)
    if flagged:
        log.warning(f"⚠️  {len(flagged)} page pairs flagged for dedup review. Fix before publishing.")
        dedup_path = output_dir / "dedup_flagged.csv"
        with open(dedup_path, "w") as f:
            f.write("slug_a,slug_b,similarity_score\n")
            for a, b, score in flagged:
                f.write(f"{a},{b},{score}\n")
        log.info(f"Dedup report saved to {dedup_path}")
        
        if args.mode == "publish":
            log.warning("Publishing paused — resolve dedup flags first. Re-run with --mode publish after fixes.")
            sys.exit(1)
    
    # Publish
    if args.mode == "publish":
        publisher = WordPressPublisher()
        published_urls = []
        results = {"published": 0, "skipped": 0, "failed": 0}
        
        for page in generated:
            result = publisher.publish_page(
                slug=page["slug"],
                title=page["title"],
                content=page["content"],
                meta_description=page["meta_description"]
            )
            if result.get("skipped"):
                results["skipped"] += 1
            elif result.get("success"):
                results["published"] += 1
                published_urls.append(page["url"])
            else:
                results["failed"] += 1
            
            time.sleep(PUBLISH_DELAY)
        
        log.info(f"\n{'='*50}")
        log.info(f"PUBLISH COMPLETE:")
        log.info(f"  Published: {results['published']}")
        log.info(f"  Skipped (already exist): {results['skipped']}")
        log.info(f"  Failed: {results['failed']}")
        
        if published_urls:
            publisher.submit_to_gsc(published_urls)
    
    else:
        log.info(f"\n✓ Wave {args.wave} — {len(generated)} pages generated. Run with --mode publish to deploy.")


if __name__ == "__main__":
    main()